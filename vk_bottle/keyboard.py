from vkbottle import Keyboard, KeyboardButtonColor, Text, OpenLink, Location, EMPTY_KEYBOARD
from vkbottle.bot import Message
from vk_bottle.api import bot
from openpyxl import load_workbook
import openpyxl

@bot.on.private_message(text="Вернуться назад")
async def back_menu(message: Message):
    users_info = await bot.api.users.get(message.from_id)
    keyboard = (
        Keyboard(one_time=False, inline=False)
        .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
        .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
        .row()
        .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
    )
    keyboard_menu = await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)
    return keyboard_menu

@bot.on.private_message(text="Здравствуйте! Меня заинтересовал этот товар.")
async def buy_osn_menu(message: Message):
    user_id = message.from_id
    with open('users.txt', 'w')as f:
        f.write(f'{user_id}\n')
    with open('users.txt', 'r') as f:
        a = f.read().split()
    b = []
    with open('user_bd.txt', 'r') as f:
        s = f.read().split()
    for i in range(0, len(a)):
        for j in range(0, len(s)):
            if a[i] == s[j]:
                break
        else:
            b.append(a[i])
    with open('user_bd.txt', 'a') as f:
        for i in range(0, len(b)):
            f.write(f'\n{b[i]}')
    with open('user_bd.txt', 'r') as f:
        a = f.read().split('\n')

    wb = openpyxl.Workbook()
    for i in range(len(a)):
        users_info = await bot.api.users.get(a[i])

        wb.create_sheet(title='Лист1', index=0)
        sheet = wb['Лист1']

        sheet[f'A1'] = 'id'
        sheet[f'B1'] = 'first_name'
        sheet[f'C1'] = 'last_name'

        sheet[f'A{i+2}'] = '{}'.format(users_info[0].id)
        sheet[f'B{i+2}'] = '{}'.format(users_info[0].first_name)
        sheet[f'C{i+2}'] = '{}'.format(users_info[0].last_name)
    sheet1 = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(sheet1)
    sheet1 = wb.get_sheet_by_name('Лист11')
    wb.remove_sheet(sheet1)
    wb.save('user_bd.xlsx')
    users_info_osn = await bot.api.users.get(message.from_id)

    keyboard = (
        Keyboard(one_time=False, inline=False)
        .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
        .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
        .row()
        .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
    )

    keyboard_menu = await message.answer("Рад тебя здесь видеть, {}! Если Вас заинтересовал данный товар, то Вы можете более подробно "
                         "узнать о нем в разделе\"Виды ботов в продаже\". Если же Вам представленной информации мало, "
                         "то вызвите нашего консултанта, выбрав \"Вызов консультанта\"".format(users_info_osn[0].first_name), keyboard=keyboard)
    return keyboard_menu
