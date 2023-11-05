from vkbottle import API
from vkbottle.bot import Bot
from vkbottle.bot import Message
from vkbottle import Keyboard, KeyboardButtonColor, Text, OpenLink, Location, EMPTY_KEYBOARD
from vkbottle import CtxStorage, BaseStateGroup
from vkbottle import PhotoMessageUploader
import requests
import pathlib
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import pyqiwi
import time
import gc

api = API("Key")
bot = Bot(api=api)

ctx = CtxStorage()

try:
    wallet = pyqiwi.Wallet(token='daa1188f358048cf57e040b61cee35fa', number='89047916581')
except Exception:
    print("Ошибка с сервером")

@bot.on.private_message(text='Баланс', peer_ids=388083972)
async def balance(message: Message):
    await message.answer(f'Ваш баланс: {wallet.balance(643)} руб.')
    del message.text


class EDITOR(BaseStateGroup):
    DATABASE = 0
    APP = 1
    ID = 2
    DEL = 3
    ABC = 4
    NEW = 5
    END = 6




@bot.on.private_message(lev="Редактировать базу данных", peer_ids=388083972)
async def first(message: Message):
    keyboard = (
        Keyboard(one_time=False, inline=False)
        .add(Text('Добваить новый столбец'), color=KeyboardButtonColor.PRIMARY)
        .row()
        .add(Text('Удалить столбец'), color=KeyboardButtonColor.PRIMARY)
        .row()
        .add(Text('Редоктировать пользователя'), color=KeyboardButtonColor.PRIMARY)
        .row()
        .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
    )
    await message.answer("Выберите действие",
                         keyboard=keyboard)
    await bot.state_dispenser.set(message.peer_id, EDITOR.DATABASE)
    del message.text


@bot.on.private_message(state=EDITOR.DATABASE)
async def editor(message: Message):
    if message.text == 'Добваить новый столбец':
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer('Введите название столбца', keyboard=keyboard)
        await bot.state_dispenser.set(message.peer_id, EDITOR.APP)
        del message.text
    elif message.text == 'Редоктировать пользователя':
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer('Введите id пользователя, чьи данные хотите изменить', keyboard=keyboard)
        await bot.state_dispenser.set(message.peer_id, EDITOR.ID)
        del message.text
    elif message.text == 'Удалить столбец':
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer('Введите название столбца, который хотите удалить', keyboard=keyboard)
        await bot.state_dispenser.set(message.peer_id, EDITOR.DEL)
        del message.text
    elif message.text == 'Вернуться назад':
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )

        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)
        del message.text


@bot.on.private_message(state=EDITOR.DEL)
async def id(message: Message):
    df = pd.read_excel('user_bd.xlsx')
    name_column = []
    for i in df.columns:
        name_column.append(i)
    if message.text in name_column:
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        wb = load_workbook('/root/user_bd.xlsx')
        sheet = wb.active
        column = 0
        for i in range(len(name_column)):
            if message.text == name_column[i]:
                column = i + 1
                break
        sheet.delete_cols(column, 1)
        wb.save('user_bd.xlsx')
        await message.answer("Столбец удален", keyboard=keyboard)
        await bot.state_dispenser.delete(message.peer_id)
        del message.text
    else:
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer("Такое столбец уже существует", keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)
        del message.text


@bot.on.private_message(state=EDITOR.ID)
async def id(message: Message):
    u_id = []
    wb = load_workbook('user_bd.xlsx')
    sheet = wb.active
    level = sheet.max_row

    for i in range(2, level + 1):
        k = sheet[f'A{i}'].value
        u_id.append(f'{k}')

    if message.text in u_id:

        ctx.set('id', message.text)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        wb = load_workbook('user_bd.xlsx')
        sheet = wb.active
        row = 0

        for j in range(0, level - 1):
            if str(message.text) == u_id[j]:
                row = sheet[f'A{j + 2}'].row

                break
        ctx.set('row', row)
        b = []
        df = pd.read_excel('user_bd.xlsx')
        name_column = []
        for i in df.columns:
            name_column.append(i)

        ABC = []
        for g in range(len(name_column)):
            index_no = df.columns.get_loc(name_column[g])
            last_column_letter = get_column_letter(index_no + 1)
            ABC.append(last_column_letter)

        for i in range(0, len(ABC)):
            b.append(sheet[f'{ABC[i]}1'].value)
        b.append('\n')
        for i in range(len(ABC)):
            a = sheet[f'{ABC[i]}{row}'].value

            b.append(str(a))

        w = ' '.join(b)
        await message.answer(f'{w}', keyboard=keyboard)
        await message.answer('Выберите название столбца, чьи данные хотите изменить')
        await bot.state_dispenser.set(message.peer_id, EDITOR.ABC)
        del message.text

    elif message.text == 'Вернуться назад':
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )

        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)
        del message.text
    else:
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )

        await message.answer("Данного пользователя нет в базе данных", keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)


@bot.on.private_message(state=EDITOR.APP)
async def app(message: Message):
    data = pd.read_excel('user_bd.xlsx')
    name_column = []
    for i in data.columns:
        name_column.append(i)

    wb = load_workbook('user_bd.xlsx')
    sheet = wb.active
    l = sheet.max_column

    if message.text not in name_column:
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        name = sheet.cell(row=1, column=l + 1)
        name.value = message.text
        df = pd.read_excel('user_bd.xlsx')
        df[name.value] = '0'
        df.to_excel('user_bd.xlsx')
        wl = load_workbook('user_bd.xlsx')
        sheet1 = wl.active
        sheet1.delete_cols(1, 1)
        wl.save('user_bd.xlsx')
        await message.answer("Новый столбец добавлен", keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)

    elif message.text == 'Вернуться назад':
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )

        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)
        del message.text
    else:
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer("Такое столбец уже существует", keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)
        del message.text


@bot.on.private_message(state=EDITOR.ABC)
async def app(message: Message):
    data = pd.read_excel('user_bd.xlsx')
    name_column = []
    for i in data.columns:
        name_column.append(i)

    if message.text in name_column:
        ctx.set('column', message.text)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer('Введите новые данные', keyboard=keyboard)
        await bot.state_dispenser.set(message.peer_id, EDITOR.NEW)
        del message.text
    elif message.text == 'Вернуться назад':
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )

        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)

    else:
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer("Такого столбеца не существует", keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)
        del message.text


@bot.on.private_message(state=EDITOR.NEW)
async def app(message: Message):
    if message.text == 'Вернуться назад':
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )

        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)
        del message.text
    else:
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        df = pd.read_excel('user_bd.xlsx')
        index_no = df.columns.get_loc(ctx.get('column'))
        last_column_letter = get_column_letter(index_no + 1)
        wb = load_workbook('user_bd.xlsx')
        sheet = wb.active
        sheet[f"{last_column_letter}{ctx.get('row')}"].value = message.text
        wb.save('user_bd.xlsx')
        ctx.delete('row')
        await message.answer("Данные измененны", keyboard=keyboard)
        await bot.state_dispenser.delete(message.peer_id)
        del message.text


class BD(BaseStateGroup):
    DATABASE = 0
    FIRST_NAME = 1
    LAST_NAME = 2
    END = 3


@bot.on.private_message(lev="База данных", peer_ids=388083972)
async def us_bd(message: Message):
    keyboard = (
        Keyboard(one_time=True, inline=False)
        .add(Text("Вывод всей базы данных"), color=KeyboardButtonColor.PRIMARY)
        .row()
        .add(Text("Поиск данных клиента"), color=KeyboardButtonColor.PRIMARY)
        .row()
        .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
    )
    await message.answer("Выберите действие",
                         keyboard=keyboard)
    await bot.state_dispenser.set(message.peer_id, BD.DATABASE)
    del message.text


@bot.on.private_message(state=BD.DATABASE)
async def look(message: Message):
    if message.text == "Вывод всей базы данных":
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        wb = load_workbook('user_bd.xlsx')
        sheet = wb.active
        level = sheet.max_row
        b = []
        for i in range(1, level):
            po = [sheet[f'A{i + 1}'].value, sheet[f'B{i + 1}'].value, sheet[f'C{i + 1}'].value]
            b.append(po)
        e = [str(item) for sublist in b for item in sublist]

        k = 0
        for i in range(3, len(e), 3):
            e.insert(i + k, '\n')
            k += 1

        w = ' '.join(e)

        await message.answer(f'{w}', keyboard=keyboard)
        await bot.state_dispenser.delete(message.peer_id)
        del message.text
    elif message.text == "Поиск данных клиента":
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer('Введите имя пользователя', keyboard=keyboard)
        await bot.state_dispenser.set(message.peer_id, BD.FIRST_NAME)
        del message.text
    elif message.text == 'Вернуться назад':
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )

        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)
        del message.text


@bot.on.private_message(state=BD.FIRST_NAME)
async def f_name(message: Message):
    name = []
    wb = load_workbook('user_bd.xlsx')
    sheet = wb.active
    level = sheet.max_row
    l = sheet.max_column
    for i in range(1, level):
        po = sheet[f'B{i + 1}'].value
        name.append(po)
    if message.text in name:

        b = []
        data = pd.read_excel('user_bd.xlsx')
        name_column = []
        bd = []
        for i in data.columns:
            name_column.append(i)

        for i in range(1, level):
            if message.text == sheet[f'B{i + 1}'].value:
                b.append('https://vk.com/id' + f'{sheet[f"A{i + 1}"].value}')
                for j in range(2, l + 1):
                    e = sheet.cell(row=i + 1, column=j)

                    b.append(str(e.value))
        for i in range(0, len(b), l):
            bd.append(b[i:i + l])

        df = pd.DataFrame(bd, columns=name_column)

        df.to_excel('name.xlsx')
        wb = load_workbook('name.xlsx')
        sheet = wb.active
        sheet.delete_cols(1, 1)
        wb.save('name.xlsx')
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer('Введите фамилию пользователя', keyboard=keyboard)
        await bot.state_dispenser.set(message.peer_id, BD.LAST_NAME)
        del message.text
    elif message.text == 'Вернуться назад':
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )
        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)
        del message.text

    elif message.text not in name:
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer("Данного имени нет в базе данных",
                             keyboard=keyboard)
        await bot.state_dispenser.delete(message.peer_id)
        del message.text


@bot.on.private_message(state=BD.LAST_NAME)
async def l_name(message: Message):
    surname = []
    wb = load_workbook('name.xlsx')
    sheet = wb.active
    k = sheet.max_row

    l = sheet.max_column

    for i in range(1, k):
        po = sheet[f'C{i + 1}'].value
        surname.append(po)
    print(surname)
    if message.text in surname:

        b = []

        for i in range(1, k):
            if message.text == sheet[f'C{i + 1}'].value:
                b.append(sheet[f'A{i + 1}'].value)
                for j in range(2, l + 1):
                    e = sheet.cell(row=i + 1, column=j)
                    b.append(e.value)
        print(b)
        wb.save('name.xlsx')
        for i in range(l, len(b), l):
            b.insert(i, '\n')
        print(b)
        w = ' '.join(b)
        print(w)
        await message.answer(f'{w}')
        await bot.state_dispenser.delete(message.peer_id)
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )
        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)
        del message.text

    elif message.text == 'Вернуться назад':
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )
        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)

        await bot.state_dispenser.delete(message.peer_id)
        del message.text
    elif message.text not in surname:
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
        )
        await message.answer("Данного пользователя с такими именем и фамилией нет в базе данных",
                             keyboard=keyboard)
        await bot.state_dispenser.delete(message.peer_id)
        del message.text


class Buy(BaseStateGroup):
    NAME = 0
    BUY = 1
    FINISH = 2
    ERRORBUY = 3
    END = 4


@bot.on.private_message(lev="Заказать бота")
async def buy_name(message: Message):
    keyboard = (
        Keyboard(one_time=True, inline=False)
        .add(Text("Торговец"), color=KeyboardButtonColor.PRIMARY)
        .row()
        .add(Text("Автоответчик"), color=KeyboardButtonColor.PRIMARY)
        .row()
        .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
    )
    await message.answer("Введите название бота, которого хотите заказать",
                         keyboard=keyboard)
    await bot.state_dispenser.set(message.peer_id, Buy.NAME)
    del message.text


@bot.on.private_message(state=Buy.NAME)
async def buy_bay(message: Message):
    try:
        if message.text == "Вернуться назад":
            users_info = await bot.api.users.get(message.from_id)
            keyboard = (
                Keyboard(one_time=False, inline=False)
                .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
                .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
                .row()
                .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
            )
            await bot.state_dispenser.delete(message.peer_id)
            await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)
            del message.text

        elif message.text == 'Автоответчик':
            ctx.set('name', message.text)
            paypal = wallet.balance(643)
            ctx.set('paypal', paypal)
            ctx.set('sum', 10.00)
            pay = pyqiwi.generate_form_link(99, '79047916581', ctx.get('sum'), ctx.get('name'))

            keyboard = (
                Keyboard(one_time=False, inline=False)
                .add(Text("Оплатил"), color=KeyboardButtonColor.POSITIVE)
                .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
            )
            await message.answer(pay, keyboard=keyboard)
            await bot.state_dispenser.set(message.peer_id, Buy.BUY)
            del message.text

        elif message.text == 'Торговец':
            ctx.set('name', message.text)
            paypal = wallet.balance(643)
            ctx.set('paypal', paypal)
            ctx.set('sum', 11.00)
            pay = pyqiwi.generate_form_link(99, '79047916581', ctx.get('sum'), ctx.get('name'))
            keyboard = (
                Keyboard(one_time=False, inline=False)
                .add(Text("Оплатил"), color=KeyboardButtonColor.POSITIVE)
                .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
            )
            await message.answer(pay, keyboard=keyboard)
            await bot.state_dispenser.set(message.peer_id, Buy.BUY)
            del message.text
    except Exception:
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )
        await bot.state_dispenser.delete(message.peer_id)
        await message.answer(
            "Извините, {}, но оплата временно не доступна из-за большого количества запросов на оплату. Повторите попытку через минут 15-20 или свяжитесь с консультаном для заказа через живого человека.".format(
                users_info[0].first_name), keyboard=keyboard)
        del message.text


@bot.on.private_message(state=Buy.BUY)
async def end_buy(message: Message):
    if message.text == "Вернуться назад":
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )
        ctx.delete('sum')
        ctx.delete('name')
        ctx.delete('paypal')
        await bot.state_dispenser.delete(message.peer_id)
        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)
        del message.text
    elif message.text == 'Оплатил':
        k = 0
        await message.answer('Проверка оплаты длиться в течение минуты')
        while k != 12:
            time.sleep(5)
            try:
                if ctx.get('paypal') != wallet.balance(643):
                    break
            except Exception:
                keyboard = (
                    Keyboard(one_time=False, inline=False)
                    .add(Text("ОЩИБКА"), color=KeyboardButtonColor.PRIMARY)
                    .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
                )
                await message.answer(
                    f"Вы заказали: {ctx.get('name')} на сумму {ctx.get('paypal')}, но из-за сбоев на севрере мы не смогли проверить на подлинность вашу оплату. Если вы правда оплатили заказ, то нажмите \"ОШИБКА\", и ждите консультанта. В ином случае \"Вернутсья назад\".",
                    keyboard=keyboard)

                await bot.state_dispenser.set(message.peer_id, Buy.ERRORBUY)
                del message.text
            k += 1
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )
        if ctx.get('paypal') + ctx.get('sum') == wallet.balance(643):

            await message.answer(f"Спасибо за заказ!\n\nВы заказали: {ctx.get('name')} на сумму {ctx.get('paypal')}",
                                 keyboard=keyboard)
            await bot.api.messages.send(peer_id=388083972, random_id=0,
                                        message=f"Заказа: {ctx.get('name')} на сумму {ctx.get('sum')}\n\nОт https://vk.com/{message.peer_id}")
            await bot.state_dispenser.delete(message.peer_id)
            del message.text
            ctx.delete('sum')
            ctx.delete('name')
            ctx.delete('paypal')
        elif ctx.get('paypal') + ctx.get('sum') != wallet.balance(643) and ctx.get('paypal') != wallet.balance(643):

            await message.answer(
                "Вы не оплатили полную стоимость товара или переплатили. В ближайшее время с важи свяжется наш менеджер для решения проблемы.",
                keyboard=keyboard)
            await bot.api.messages.send(peer_id=388083972, random_id=0,
                                        message=f"Ошибка в заказе: {ctx.get('name')} на сумму {ctx.get('paypal')}\n\nОт https://vk.com/{message.peer_id}")
            ctx.delete('sum')
            ctx.delete('name')
            ctx.delete('paypal')
            await bot.state_dispenser.delete(message.peer_id)
            del message.text
        elif ctx.get('paypal') == wallet.balance(643):
            keyboard = (
                Keyboard(one_time=False, inline=False)
                .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
            )
            await message.answer(
                "Оплата не была проведена. Повторите попытку или свяжитесь с консультаном, если с вашего счета списались деньги, но вышло это сообщение.",
                keyboard=keyboard)
            ctx.delete('sum')
            ctx.delete('name')
            ctx.delete('paypal')
            await bot.state_dispenser.delete(message.peer_id)
            del message.text

    else:
        await message.answer(
            'Не понял вашего ответа. Отправьте \"Оплатил\" для подтверждения оплаты или \"Вернуться назад\", в случае, если вы не собираетесь платить.')
        ctx.delete('sum')
        ctx.delete('name')
        ctx.delete('paypal')
        await bot.state_dispenser.set(message.peer_id, Buy.BUY)
        del message.text


@bot.on.private_message(state=Buy.ERRORBUY)
async def error_buy(message: Message):
    if message.test == 'ОШИБКА':
        await bot.api.messages.send(peer_id=388083972, random_id=0,
                                    message=f"!!!ОШИБКА!!!\n\n\nЗаказа: {ctx.get('name')} на сумму {ctx.get('paypal')}\n\nОт https://vk.com/{message.peer_id}")
        await message.answer('Мы оповестили нашего сотрудника об ошибке. В ближайшее время он с вами свяжется')
        ctx.delete('sum')
        ctx.delete('name')
        ctx.delete('paypal')
        await bot.state_dispenser.delete(message.peer_id)
        del message.text
    else:
        users_info = await bot.api.users.get(message.from_id)
        keyboard = (
            Keyboard(one_time=False, inline=False)
            .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
            .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
            .row()
            .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
        )
        ctx.delete('sum')
        ctx.delete('name')
        ctx.delete('paypal')
        await bot.state_dispenser.delete(message.peer_id)
        await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)
        del message.text


class Ras(BaseStateGroup):
    TEXT = 0
    PHOTO = 1
    END = 2


@bot.on.private_message(lev=["рассылка", "Рассылка"], peer_ids=388083972)
async def ras_text(message: Message):
    await bot.state_dispenser.set(message.peer_id, Ras.TEXT)
    del message.text
    return "Введите текс рассылки"


@bot.on.private_message(state=Ras.TEXT)
async def ras_photo(message: Message):
    ctx.set('text', message.text)
    await bot.state_dispenser.set(message.peer_id, Ras.PHOTO)
    del message.text
    return "Пришли фото, если надо, или отправьте \"СТОП\", если без фотографии"


@bot.on.private_message(state=Ras.PHOTO)
async def end_ras(message: Message):
    if message.text == 'СТОП':
        wb = load_workbook('user_bd.xlsx')
        sheet = wb.active
        l = sheet.max_row
        bd = []
        for i in range(2, l + 1):
            bd.append(sheet[f'A{i}'].value)
        for user_id in bd:
            await bot.api.messages.send(peer_id=int(user_id), random_id=0, message=f"{ctx.get('text')}")
        ctx.delete('text')
        await bot.state_dispenser.delete(message.peer_id)
        del message.text
        return "Рассылка отправленна"

    else:
        Path(f'files/{message.peer_id}/').mkdir(parents=True, exist_ok=True)
        url = message.attachments[0].photo.sizes[-5].url

        src = f'files/{message.peer_id}/'
        with open(src + 'rassilka.png', 'wb') as f:
            f.write(requests.get(url).content)
        photo_upd = PhotoMessageUploader(bot.api)
        photo = await photo_upd.upload(f"files/{message.peer_id}/rassilka.png")

        wb = load_workbook('user_bd.xlsx')
        sheet = wb.active
        i = []
        l = sheet.max_row
        for j in range(2, l + 1):
            i.append(sheet[f'A{j}'].value)
        for user_id in i:
            await bot.api.messages.send(peer_id=int(user_id), random_id=0, message=f"{ctx.get('text')}\n",
                                        attachment=photo)
        ctx.delete('text')
        file = pathlib.Path(f"files/{message.peer_id}/rassilka.png")
        file.unlink()
        await bot.state_dispenser.delete(message.peer_id)
        del message.text
        return " Рассылка отправленна"


@bot.on.private_message(text="Здравствуйте! Меня заинтересовал этот товар.")
async def zhandler(message: Message):
    user_id = message.from_id
    users_info = await bot.api.users.get(user_id)

    wb = load_workbook('user_bd.xlsx')
    sheet = wb.active
    level = sheet.max_row
    df = pd.read_excel('user_bd.xlsx')
    name_column = []
    for i in df.columns:
        name_column.append(i)
    ABC = []
    for g in range(len(name_column)):
        index_no = df.columns.get_loc(name_column[g])
        last_column_letter = get_column_letter(index_no + 1)
        ABC.append(last_column_letter)

    sheet[f'A{level + 1}'] = '{}'.format(users_info[0].id)
    sheet[f'B{level + 1}'] = '{}'.format(users_info[0].first_name)
    sheet[f'C{level + 1}'] = '{}'.format(users_info[0].last_name)
    for j in range(3, len(ABC)):
        sheet[f'{ABC[j]}{level + 1}'] = '0'

    for i in range(2, level + 1):
        for j in range(i + 1, level + 2):
            if sheet[f'A{i}'].value == sheet[f'A{j}'].value:
                sheet.delete_rows(j, 1)
                break
    wb.save('user_bd.xlsx')

    keyboard = (
        Keyboard(one_time=False, inline=False)
        .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
        .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
        .row()
        .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
    )

    await message.answer("Рад тебя здесь видеть, {}! Если Вас заинтересовал данный товар, то Вы можете более подробно "
                         "узнать о нем в разделе\"Виды ботов в продаже\". Если же Вам представленной информации мало, "
                         "то вызвите нашего консултанта, выбрав \"Вызов консультанта\"".format(
        users_info[0].first_name), keyboard=keyboard)
    del message.text


@bot.on.private_message(text=["Начать", "Привет"])
async def shandler(message: Message):
    user_id = message.from_id
    users_info = await bot.api.users.get(user_id)

    wb = load_workbook('user_bd.xlsx')
    sheet = wb.active
    level = sheet.max_row
    df = pd.read_excel('user_bd.xlsx')
    name_column = []
    for i in df.columns:
        name_column.append(i)
    ABC = []
    for g in range(len(name_column)):
        index_no = df.columns.get_loc(name_column[g])
        last_column_letter = get_column_letter(index_no + 1)
        ABC.append(last_column_letter)

    sheet[f'A{level + 1}'] = '{}'.format(users_info[0].id)
    sheet[f'B{level + 1}'] = '{}'.format(users_info[0].first_name)
    sheet[f'C{level + 1}'] = '{}'.format(users_info[0].last_name)
    for j in range(3, len(ABC)):
        sheet[f'{ABC[j]}{level + 1}'] = '0'

    for i in range(2, level + 1):
        for j in range(i + 1, level + 2):
            if sheet[f'A{i}'].value == sheet[f'A{j}'].value:
                sheet.delete_rows(j, 1)
                break

    wb.save('user_bd.xlsx')

    keyboard = (
        Keyboard(one_time=False, inline=False)
        .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
        .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
        .row()
        .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
    )

    await message.answer("Рад тебя здесь видеть, {}! Чем могу помочь?".format(users_info[0].first_name),
                         keyboard=keyboard)
    del message.text


@bot.on.private_message(text="Вернуться назад")
async def bhandler(message: Message):
    users_info = await bot.api.users.get(message.from_id)
    keyboard = (
        Keyboard(one_time=False, inline=False)
        .add(Text("Как сделать заказ"), color=KeyboardButtonColor.SECONDARY)
        .add(Text("Виды ботов в продаже"), color=KeyboardButtonColor.SECONDARY)
        .row()
        .add(Text("Вызов консультанта"), color=KeyboardButtonColor.POSITIVE)
    )

    await message.answer("Чем могу еще помочь, {}?".format(users_info[0].first_name), keyboard=keyboard)
    del message.text


@bot.on.private_message(text="Как сделать заказ")
async def handler(message: Message):
    keyboard = (
        Keyboard(one_time=False, inline=False)
        .add(Text("Заказать бота"), color=KeyboardButtonColor.POSITIVE)
        .row()
        .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
    )

    await message.answer(
        'Заказать бота можно через ссылку Qiwi, отправленную вам в последующих меню После оплаты ' \
        'отправьте боту \"Заказать бота\", после чего следуйте инструкциям.', keyboard=keyboard)
    del message.text


@bot.on.private_message(text="Виды ботов в продаже")
async def vhandler(message: Message):
    keyboard = (
        Keyboard(one_time=False, inline=False)
        .add(Text("Вернуться назад"), color=KeyboardButtonColor.NEGATIVE)
    )
    await message.answer('Все доступные боты у меня:\n✔  Торговец\n✔  Автоответчик', keyboard=keyboard)
    del message.text


@bot.on.private_message(text="Вызов консультанта")
async def dzin(message: Message):
    await message.answer(
        'Я уже вызвал консультанта, так что он Вам напишет как можно быстрее. А пока посмотрите мой ассортимент')
    await bot.api.messages.send(peer_id=388083972, random_id=0,
                                message=f'Вас вызывает https://vk.com/id{message.from_id}')
    del message.text


gc.collect()
bot.run_forever()
